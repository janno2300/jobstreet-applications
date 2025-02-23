import imaplib
from itertools import chain
import email
from email.header import decode_header
import openpyxl
from openpyxl.styles import Alignment
from datetime import datetime
from datetime import timedelta
import re
from bs4 import BeautifulSoup
import quopri
import unicodedata
import configparser
import io
import sys
import os

def main():
    try:
        config = load_config()

        file_path = config.get("Settings", "file_path")
        with io.open(file_path, "r"):
            pass
        is_file_open(file_path)
            
        username = config.get("Settings", "username")
        password = config.get("Settings", "password")

        criteria = {
            "FROM": config.get("Criteria", "from_email"),
            "SINCE": config.get("Criteria", "since_date"),
        }

        uid_max = config.getint("Other", "uid_max")
        
        submitted_applications(file_path, username, password, uid_max, criteria)
        viewed_applications(file_path, username, password, uid_max, criteria)
        closed_applications(file_path, username, password, uid_max, criteria)
        
    except PermissionError as e:
        print(f"Error: Unable to access '{file_path}' due to permission issues.")
        print("Possible Solutions:")
        print("1. Close the file if it's open in Excel or another program.")
        print("2. Check file permissions and ensure you have access.")
        print("3. Try running the script as Administrator.")
        print(f"Technical Details: {e}")
        sys.exit(1)

    except Exception as e:
        print(f"Unexpected error: {e}")
        sys.exit(1)
        
def is_file_open(file_path):
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"Error: The file '{file_path}' does not exist.")

    try:
        wb = openpyxl.load_workbook(file_path)
        wb.save(file_path)
    except PermissionError as e:
        print(f"Error: Unable to access '{file_path}' due to permission issues.")
        print("Possible Solutions:")
        print("1. Close the file if it's open in Excel or another program.")
        print("2. Check file permissions and ensure you have access.")
        print("3. Try running the script as Administrator.")
        print(f"Technical Details: {e}")
        sys.exit(1)
    except Exception as e:
        print(f"Unexpected error: {e}")
        sys.exit(1)
    
def load_config(file_name="config.ini"):
    config = configparser.ConfigParser()
    config.read(file_name)
    return config

def format_email_date(email_date):
    email_date_str = email_date.split(" (")[0]
    email_date_obj = datetime.strptime(email_date_str, "%a, %d %b %Y %H:%M:%S %z")
    return email_date_obj.strftime("%Y-%m-%d")

def format_application_date(new_date):
    cleaned_date = unicodedata.normalize("NFKC", new_date)
    cleaned_date = re.sub(r'\s+', ' ', cleaned_date).strip()
    cleaned_date = re.sub(r'(\d)\s+(\d)', r'\1\2', cleaned_date)
    cleaned_date = re.sub(r'(\w)\s+(\w)', r'\1\2', cleaned_date)  
    
    has_year = re.search(r'\d{4}', cleaned_date)

    if has_year:
        formatted_date = datetime.strptime(cleaned_date, "%d %b %Y").strftime("%Y-%m-%d")
    else:
        match = re.search(r"^(.*?)\s*(?:Applicationinformation|Similarjobsyoumight)", cleaned_date, re.IGNORECASE)
        if match:
            cleaned_date = match.group(1).strip()
            # cleaned_date = re.sub(r"(\d{1,2})\s*([A-Za-z]{1})\s*([A-Za-z]{2})", r"\1 \2\3", cleaned_date)
        # else:
            # cleaned_date = re.sub(r"(\d{1,2})([A-Za-z]{3})", r"\1 \2", cleaned_date)
        cleaned_date = re.sub(r"(\d{1,2})\s*([A-Za-z]{1})\s*([A-Za-z]{2})", r"\1 \2\3", cleaned_date)
        cleaned_date = re.sub(r"\b([A-Za-z])\s+([A-Za-z])\b", r"\1\2", cleaned_date)
        cleaned_date = re.sub(r"(\d{1,2})([A-Za-z]{3})", r"\1 \2", cleaned_date)
            
        one_month_ago = datetime.today().replace(day=1) - timedelta(days=1) 
        assumed_year = one_month_ago.year
        
        full_date = f"{cleaned_date} {assumed_year}"
        formatted_date = datetime.strptime(full_date, "%d %b %Y").strftime("%Y-%m-%d")
        
    return formatted_date

def search_string(uid_max, criteria):
    c = list(map(lambda t: (t[0], '"'+str(t[1])+'"'), criteria.items())) + [('UID', '%d:*' % (uid_max+1))]
    return '(%s)' % ' '.join(chain(*c))

def submitted_applications(file_path, username, password, uid_max, criteria):
    try:
        imap_server = imaplib.IMAP4_SSL("imap.mail.yahoo.com", 993)
        imap_server.login(username, password)
        imap_server.select("INBOX")
        
        criteria['SUBJECT'] = 'Your application was successfully submitted'        
        result, data = imap_server.uid('search', None, search_string(uid_max, criteria))
        
        start_row = 3
        is_file_open(file_path)
        wb = openpyxl.load_workbook(file_path)
        ws = wb["Applications"]
        
        num_list = {0}
        
        for num in data[0].split():
            num_int = int(num.decode())
            
            if num_int not in num_list:
                result, msg_data = imap_server.uid('fetch', num, '(RFC822)')
                for response_part in msg_data:
                    if isinstance(response_part, tuple):
                        msg = email.message_from_bytes(response_part[1])
                        
                        subject = decode_header(msg["Subject"])[0][0]
                        if isinstance(subject, bytes):
                            subject = subject.decode()
                            
                        print("--------------------------------------------------------------------------------------")     
                        print(f"UID: {num_int} | Subject: {subject}")
                        
                        if subject == 'Your application was successfully submitted':
                            email_date = format_email_date(msg["Date"])
                            
                            ws.cell(row=start_row, column=1, value=num.decode())  # UID
                            ws.cell(row=start_row, column=4, value=email_date).alignment = Alignment(horizontal='center')  # Date
                            wb.save(file_path)
                            
                            if msg.is_multipart():
                                for part in msg.walk():
                                    content_type = part.get_content_type()
                                    content_disposition = str(part.get("Content-Disposition"))

                                    if content_type == "text/plain" and "attachment" not in content_disposition:
                                        raw_email_text = part.get_payload(decode=True)
                                        
                            else:
                                raw_email_text = msg.get_payload(decode=True)
                                
                            
                            if raw_email_text:
                                decoded_text = quopri.decodestring(raw_email_text).decode("utf-8")
                                decoded_text = re.sub(r"=\n", "", decoded_text)
                                decoded_text = re.sub(r"\s+", " ", decoded_text).strip()
                            else:
                                decoded_text = ""
                                
                            match = re.search(
                                r"Your application for (.*?) was successfully submitted to (.*?)\. Each",
                                decoded_text,
                                re.DOTALL
                            )
                            
                            position = ''
                            company = ''
                            job_link = ''
                            location = ''

                            if match:
                                position = match.group(1).strip()
                                company = match.group(2).strip()
                                
                                position = re.sub(r"\s+", " ", position)
                                company = re.sub(r"\s+", " ", re.sub(r"\.\.$", ".", company)).strip()
                                
                            # print(f"Extracted Position: {position}")
                            # print(f"Extracted Company: {company}")
                            
                            # pattern = rf"{re.escape(position)}\s*\[\s*(https?://[^\]]+)\s*\]\s*{re.escape(company)}\s*(.*?)$"
                            pattern = rf"{re.escape(position)}\s*\[\s*(https?://[^\]]+)\s*\]\s*{re.escape(company)}\s*([\w\s,-]+)"
                            match = re.search(pattern, decoded_text, re.MULTILINE)

                            if match and company != '':
                                job_link = match.group(1).strip()
                                location = match.group(2).strip()
                                location = re.sub(r"[^\w\s,.-]", "", location).strip()
                                
                            print(f"Date: {email_date}")
                            print(f"Position: {position}")
                            print(f"Company: {company}")
                            print(f"Location: {location}")
                            
                            ws.cell(row=start_row, column=2, value=company)  # Company
                            ws.cell(row=start_row, column=3, value=position)  # Position
                            ws.cell(row=start_row, column=8, value=location)  # Location
                            ws.cell(row=start_row, column=9, value=job_link)  # Link
                            wb.save(file_path)
                            start_row += 1
            
        wb.save(file_path)
        imap_server.logout()
            
    except PermissionError as e:
        print(f"Error: Unable to access '{file_path}' due to permission issues.")
        print("Possible Solutions:")
        print("1. Close the file if it's open in Excel or another program.")
        print("2. Check file permissions and ensure you have access.")
        print("3. Try running the script as Administrator.")
        print(f"Technical Details: {e}")
        sys.exit(1)

    except Exception as e:
        print(f"Unexpected error: {e}")
        sys.exit(1)
        
def viewed_applications(file_path, username, password, uid_max, criteria):
    try:
        imap_server = imaplib.IMAP4_SSL("imap.mail.yahoo.com", 993)
        imap_server.login(username, password)
        imap_server.select("INBOX")
        
        criteria['SUBJECT'] = 'has viewed your application for'        
        result, data = imap_server.uid('search', None, search_string(uid_max, criteria))
        
        start_row = 3
        is_file_open(file_path)
        wb = openpyxl.load_workbook(file_path)
        ws = wb["Applications"]
        
        num_list = {0}
        
        for num in data[0].split():
            num_int = int(num.decode())
            if num_int not in num_list:
                result, msg_data = imap_server.uid('fetch', num, '(RFC822)')
                for response_part in msg_data:
                    if isinstance(response_part, tuple):
                        msg = email.message_from_bytes(response_part[1])
                        
                        subject = decode_header(msg["Subject"])[0][0]
                        if isinstance(subject, bytes):
                            subject = subject.decode()
                        
                        subject = re.sub(r"\s+", " ", subject)
                            
                        print("--------------------------------------------------------------------------------------")     
                        print(f"UID: {num_int} | Subject: {subject}")
                        
                        position_search = ''
                        company_search = ''
                        application_date_search = ''
                        row_number = ''
                        
                        email_date = format_email_date(msg["Date"])
                        
                        if msg.is_multipart():
                            for part in msg.walk():
                                content_type = part.get_content_type()
                                content_disposition = str(part.get("Content-Disposition"))

                                if content_type == "text/plain" and "attachment" not in content_disposition:
                                    raw_email_text = part.get_payload(decode=True)
                                    
                        else:
                            raw_email_text = msg.get_payload(decode=True)
                            
                        
                        if raw_email_text:
                            decoded_text = quopri.decodestring(raw_email_text).decode("utf-8")
                            decoded_text = re.sub(r"=\n", "", decoded_text)
                            decoded_text = re.sub(r"\s+", " ", decoded_text).strip()
                        else:
                            decoded_text = ""
                            
                        match = re.search(
                            r"Your application for (.*?) was viewed by (.*?)\. Each",
                            decoded_text,
                            re.DOTALL
                        )
                        
                        if match:
                            position_search = match.group(1).strip()
                            company_search = match.group(2).strip()
                            
                            position_search = re.sub(r"\s+", " ", position_search)
                            company_search = re.sub(r"\s+", " ", re.sub(r"\.\.$", ".", company_search)).strip()
                            
                        # match = re.search(r'Applied on\s+([\d]{1,2} [A-Za-z]+ \d{4})', decoded_text)
                        match = re.search(r'Applied on\s+([\d\s]{1,4}[A-Za-z\s]+)(\d{4})?', decoded_text)
                        
                        if match:
                            application_date_search = match.group(1).strip()
                            application_date_search = format_application_date(application_date_search)
                        
                        matching_row = ''
                        for row in range(2, ws.max_row + 1):
                            company = ws.cell(row=row, column=2).value
                            application_date = ws.cell(row=row, column=4).value
                            position = ws.cell(row=row, column=3).value
                            
                            if isinstance(application_date, str):
                                application_date = application_date.strip()
                                
                            if company == company_search and str(application_date) == application_date_search and position == position_search:
                                matching_row = row
                                break
                                
                        if matching_row:
                            ws.cell(row=matching_row, column=5, value=email_date).alignment = Alignment(horizontal='center')  # Viewed date
                            wb.save(file_path)
                            
                        print(f"Date: {email_date}")
                        print(f"Position: {position}")
                        print(f"Company: {company}")
                        print(f"Application date: {application_date_search}")
                            
            imap_server.logout()
            
    except PermissionError as e:
        print(f"Error: Unable to access '{file_path}' due to permission issues.")
        print("Possible Solutions:")
        print("1. Close the file if it's open in Excel or another program.")
        print("2. Check file permissions and ensure you have access.")
        print("3. Try running the script as Administrator.")
        print(f"Technical Details: {e}")
        sys.exit(1)
        
    except Exception as e:
        raise Exception("Error while connecting to the Yahoo email server.") from e
        sys.exit(1)
        
def closed_applications(file_path, username, password, uid_max, criteria):
    try:
        imap_server = imaplib.IMAP4_SSL("imap.mail.yahoo.com", 993)
        imap_server.login(username, password)
        imap_server.select("INBOX")
        
        criteria['SUBJECT'] = 'has closed'        
        result, data = imap_server.uid('search', None, search_string(uid_max, criteria))
        
        start_row = 3
        is_file_open(file_path)
        wb = openpyxl.load_workbook(file_path)
        ws = wb["Applications"]
        
        num_list = {0}
        
        for num in data[0].split():
            num_int = int(num.decode())
            if num_int not in num_list:
                result, msg_data = imap_server.uid('fetch', num, '(RFC822)')
                for response_part in msg_data:
                    if isinstance(response_part, tuple):
                        msg = email.message_from_bytes(response_part[1])
                        # print(msg)
                        
                        subject = decode_header(msg["Subject"])[0][0]
                        if isinstance(subject, bytes):
                            subject = subject.decode()
                        
                        subject = re.sub(r"\s+", " ", subject)
                            
                        print("--------------------------------------------------------------------------------------")     
                        print(f"UID: {num_int} | Subject: {subject}")
                        
                        position_search = ''
                        company_search = ''
                        application_date_search = ''
                        row_number = ''
                        
                        email_date = format_email_date(msg["Date"])
                        
                        if msg.is_multipart():
                            for part in msg.walk():
                                content_type = part.get_content_type()
                                content_disposition = str(part.get("Content-Disposition"))

                                if content_type == "text/plain" and "attachment" not in content_disposition:
                                    raw_email_text = part.get_payload(decode=True)
                                    
                        else:
                            raw_email_text = msg.get_payload(decode=True)
                            
                        
                        if raw_email_text:
                            decoded_text = quopri.decodestring(raw_email_text).decode("utf-8")
                            decoded_text = re.sub(r"=\n", "", decoded_text)
                            decoded_text = re.sub(r"\s+", " ", decoded_text).strip()
                        else:
                            decoded_text = ""
                            
                        match = re.search(
                            r"the (.*?) job you applied for at (.*?)\ has expired",
                            decoded_text,
                            re.DOTALL
                        )
                        
                        if match:
                            position_search = match.group(1).strip()
                            company_search = match.group(2).strip()
                            
                            position_search = re.sub(r"\s+", " ", position_search)
                            company_search = re.sub(r"\s+", " ", re.sub(r"\.\.$", ".", company_search)).strip()
                            
                        match = re.search(r'Applied on\s+([\d\s]{1,4}[A-Za-z\s]+)(\d{4})?', decoded_text)
                        
                        if match:
                            application_date_search = match.group(1).strip()
                            application_date_search = format_application_date(application_date_search)
                            
                        match = re.search(r"Application information\s*\[.*?\]\s*(\d+)\s+candidates applied", decoded_text, re.DOTALL)

                        if match:
                            applicant_count = int(match.group(1))
                            print(f"Applicant count: {applicant_count}")
                        # else:
                            # print(f"wala")
                        
                        matching_row = ''
                        for row in range(2, ws.max_row + 1):
                            company = ws.cell(row=row, column=2).value
                            application_date = ws.cell(row=row, column=4).value
                            position = ws.cell(row=row, column=3).value
                            
                            if isinstance(application_date, str):
                                application_date = application_date.strip()
                            
                            # Consider possibility of companies changing their name
                            if company == company_search and str(application_date) == str(application_date_search) and position == position_search:
                                matching_row = row
                                break
                            elif str(application_date) == str(application_date_search) and position == position_search:
                                matching_row = row
                                break
                            elif company == company_search and position == position_search:
                                matching_row = row
                                break
                            # if company == company_search:
                            # if str(application_date) == str(application_date_search) and position == position_search:
                                # matching_row = row
                                # break
                                
                        if matching_row:
                            ws.cell(row=matching_row, column=6, value=email_date).alignment = Alignment(horizontal='center')  # Closed date
                            ws.cell(row=matching_row, column=7, value=applicant_count).alignment = Alignment(horizontal='center')  # Applicants
                            wb.save(file_path)
                            
                        print(f"Date: {email_date}")
                        print(f"Position: {position_search}")
                        print(f"Company: {company_search}")
                        # print(f"Application date 1: {application_date}")
                        print(f"Application date: {application_date_search}")
                        
        imap_server.logout()
        
    except PermissionError as e:
        print(f"Error: Unable to access '{file_path}' due to permission issues.")
        print("Possible Solutions:")
        print("1. Close the file if it's open in Excel or another program.")
        print("2. Check file permissions and ensure you have access.")
        print("3. Try running the script as Administrator.")
        print(f"Technical Details: {e}")
        sys.exit(1)
        
    except Exception as e:
        raise Exception("Error while connecting to the Yahoo email server.") from e
        sys.exit(1)

if __name__ == "__main__":
    main()